from io import BytesIO
from sqlalchemy.orm import Session
from models import FormSubmission
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

class ExportService:
    @staticmethod
    def export_to_excel(db: Session):
        """
        Export all form submissions to Excel
        
        Args:
            db: Database session
            
        Returns:
            BytesIO object containing Excel file
        """
        # Get all submissions from database
        submissions = db.query(FormSubmission).all()
        
        # Create new workbook and select active worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Form Submissions'
        
        # Define headers
        headers = [
            "Jina Kamili", "Jinsi", "Umri", "Barua Pepe", "Nambari Simu",
            "Ngazi ya Elimu", "Jina la Kozi", "Mwaka wa Kuhitimu",
            "Taasisi ya Mwisho", "Ujuzi wa Kompyuta", "Lugha", "Msaada",
            "Tarehe ya Usajili"
        ]
        
        # Write headers to the first row
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)
        
        # Write data
        for row_num, submission in enumerate(submissions, 2):  # Start from row 2
            worksheet.cell(row=row_num, column=1, value=submission.jina_kamili)
            worksheet.cell(row=row_num, column=2, value=submission.jinsi)
            worksheet.cell(row=row_num, column=3, value=submission.umri)
            worksheet.cell(row=row_num, column=4, value=submission.barua_pepe)
            worksheet.cell(row=row_num, column=5, value=submission.nambari_simu)
            worksheet.cell(row=row_num, column=6, value=submission.ngazi_elimi)
            worksheet.cell(row=row_num, column=7, value=submission.jina_kozi)
            worksheet.cell(row=row_num, column=8, value=submission.mwaka_kuhitimu)
            worksheet.cell(row=row_num, column=9, value=submission.taasisi_mwisho or "")
            worksheet.cell(row=row_num, column=10, value=submission.ujuzi_kompyuta)
            worksheet.cell(row=row_num, column=11, value=submission.lugha)
            worksheet.cell(row=row_num, column=12, value=submission.msaada)
            worksheet.cell(row=row_num, column=13, value=submission.created_at)
        
        # Adjust column widths
        for col_num, _ in enumerate(headers, 1):
            # Get maximum length of data in each column
            max_length = 0
            for row_num in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set column width based on max length (with some padding)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        
        # Create BytesIO buffer to hold Excel file
        output = BytesIO()
        
        # Save workbook to the buffer
        workbook.save(output)
        
        # Reset buffer position to beginning
        output.seek(0)
        
        return output
    
    @staticmethod
    def generate_filename():
        """Generate a filename for the Excel export"""
        now = datetime.now()
        return f"form_submissions_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
